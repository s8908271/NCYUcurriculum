# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'class.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from PyQt5.QtWidgets import QDialog, QApplication, QMessageBox, QLineEdit, QFileDialog, QDockWidget, QListWidget, QComboBox
from PyQt5.QtCore import QDir, Qt, QUrl, QObject
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
import selenium.webdriver.support.ui as ui
from lxml import html
from bs4 import BeautifulSoup as bs
from time import sleep
import os
from lxml import etree
import xlrd
import xlwt
import xlutils
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Color


class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(400, 300)
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setGeometry(QtCore.QRect(70, 60, 113, 22))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(Form)
        self.lineEdit_2.setGeometry(QtCore.QRect(70, 130, 113, 22))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(240, 90, 93, 28))
        self.pushButton.setObjectName("pushButton")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(70, 30, 55, 16))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(70, 100, 55, 16))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(30, 210, 101, 16))
        self.label_3.setObjectName("label_3")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.pushButton.setText(_translate("Form", "登入"))
        self.label.setText(_translate("Form", "帳號"))
        self.label_2.setText(_translate("Form", "密碼"))
        self.label_3.setText(_translate("Form", "請耐心等候產出"))

class AppWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.ui.lineEdit_2.setEchoMode(QLineEdit.Password)
        self.ui.pushButton.clicked.connect(self.Click)

    def Click(self):
        username = self.ui.lineEdit.text()
        password = self.ui.lineEdit_2.text()
        if (username == "" or password == ""):
            reply = QMessageBox.warning(self,"警告","帳號密碼不能為空，請輸入！")
            return
        if os.path.isfile('empty.xlsx'):
            print('exist')
        else:
            reply = QMessageBox.warning(self,"警告","檔案缺少，請重新下載")
            return

        filename = 'empty.xlsx'
        book_r = load_workbook(filename)
        sheet_1 = book_r.active

        options = Options()
        prefs = {
              'profile.default_content_setting_values' :
                  {
                  'notifications' : 2
                  }
        }
        options.add_experimental_option('prefs',prefs)
        options.add_argument('--headless')#無介面
        options.add_argument("--start-maximized")#全視窗
        options.add_argument("--incognito")#無痕
        driver = webdriver.Chrome(options=options)
        LOGIN= 'https://web085004.adm.ncyu.edu.tw/NewSite/Login.aspx?Language=zh-TW'
        driver.get(LOGIN)
        driver.find_element_by_xpath('//input[@id="TbxAccountId"]').send_keys(username)
        driver.find_element_by_xpath('//input[@id="TbxPassword"]').send_keys(password)
        driver.find_element_by_xpath('//input[@name="BtnPreLogin"]').click()
        sleep(2)
        try:
            driver.find_element_by_xpath('//*[@id="DivMessageBox"]')
            reply = QMessageBox.warning(self,"警告","帳號密碼驗證錯誤")
            return
        except:
            print('yes')
        soup = bs(driver.page_source,"lxml")
        authenticity_token = soup.find('input', {'id': 'WebPid1'}).get('value')
        LOGIN2= 'https://web085004.adm.ncyu.edu.tw/NewSite/Refer.aspx?action=/GradChk/SelGradChk.aspx&method=post&WebPid1=' + authenticity_token + '&Language=zh-TW'
        driver.get(LOGIN2)
        driver.find_element_by_xpath('//input[@name="BtnSelB"]').click()
        #driver.find_element_by_xpath('/html/body/table[4]/tbody/tr/td[3]/div[3]/form/div/input[4]').click()
        try:
            driver.find_element_by_xpath('//input[@value="當學期選課查詢"]').click()
        except:
            print('no')
        try:
            driver.find_element_by_xpath('//input[@value="選課結果查詢"]').click()
        except:
            print('no')
        soup2= bs(driver.page_source,"lxml")
        selector =etree.HTML(driver.page_source)

        try:
            selector.xpath('/html/body/table[1]/tbody/tr[2]/td[14]/text()')
            k=2
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[3]/td[14]/text()')
            k=3
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[4]/td[14]/text()')
            k=4
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[5]/td[14]/text()')
            k=5
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[6]/td[14]/text()')
            k=6
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[7]/td[14]/text()')
            k=7
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[8]/td[14]/text()')
            k=8
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[9]/td[14]/text()')
            k=9
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[10]/td[14]/text()')
            k=10
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[11]/td[14]/text()')
            k=11
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[12]/td[14]/text()')
            k=12
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[13]/td[14]/text()')
            k=13
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[14]/td[14]/text()')
            k=14
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[15]/td[14]/text()')
            k=15
            print(k)
        except:
            print('no')
        try:
            selector.xpath('/html/body/table[1]/tbody/tr[16]/td[14]/text()')
            k=16
            print(k)
        except:
            print('no')

        i =2
        while i <= k:
            try: 
#星期
                firstweek=selector.xpath('/html/body/table[1]/tbody/tr['+str(i)+']/td[14]/text()')
                firstweeks = str(firstweek)
                firstweekss = firstweeks[2:len(firstweeks)-2]

#節次
                firsttime=selector.xpath('/html/body/table[1]/tbody/tr['+str(i)+']/td[15]/text()')
                firsttimes = str(firsttime)
                firsttimess = firsttimes[2:len(firsttimes)-2]

#名稱
                firstname=selector.xpath('/html/body/table[1]/tbody/tr['+str(i)+']/td[4]/a/text()')
                firstnames = str(firstname)
                firstnamess = firstnames[2:len(firstnames)-2]

#地點
                firstwhere=selector.xpath('/html/body/table[1]/tbody/tr['+str(i)+']/td[17]/text()')
                firstwheres = str(firstwhere)
                firstwheress = firstwheres[2:len(firstwheres)-2]
        
#教師
                firstteacher=selector.xpath('/html/body/table[1]/tbody/tr['+str(i)+']/td[13]/text()')
                firstteachers = str(firstteacher)
                firstteacherss = firstteachers[2:len(firstteachers)-2]

                if len(firstweekss) > 2:
                    try:
                
            #第一個
                        if firstweekss.find('一')==0:
                           week = 7
                        pass
                        if firstweekss.find('二')==0:
                           week = 6
                        pass
                        if firstweekss.find('三')==0:
                           week = 5
                        pass
                        if firstweekss.find('四')==0:
                           week = 4
                        pass
                        if firstweekss.find('五')==0:
                           week = 3
                        pass
                        if firstweekss.find('六')==0:
                           week = 2
                        pass
                        if firstweekss.find('七')==0:
                           week = 1
                        pass
                        if firsttimess.find('1')==0:
                           time = 2
                           end = 2
                        pass
                        if firsttimess.find('2')==0:
                           time = 3
                           end = 3
                        pass
                        if firsttimess.find('3')==0:
                           time = 4
                           end = 4
                        pass
                        if firsttimess.find('4')==0:
                           time = 5
                           end = 5
                        pass
                        if firsttimess.find('F')==0:
                           time = 6
                           end = 6
                        pass
                        if firsttimess.find('5')==0:
                           time = 7
                           end = 7
                        pass
                        if firsttimess.find('6')==0:
                           time = 8
                           end = 8
                        pass
                        if firsttimess.find('7')==0:
                           time = 9
                           end = 9
                        pass
                        if firsttimess.find('8')==0:
                           time = 10
                           end = 10
                        pass
                        if firsttimess.find('9')==0:
                           time = 11
                           end = 11
                        pass
                        if firsttimess.find('A')==0:
                           time = 12
                           end = 12
                        pass
                        if firsttimess.find('B')==0:
                           time = 13
                           end = 13
                        pass
                        if firsttimess.find('C')==0:
                           time = 14
                           end = 14
                        pass
                        if firsttimess.find('D')==0:
                           time = 15
                           end = 15
                        pass
            
                        if firsttimess.find('1')==2:
                           end = 2
                        pass
                        if firsttimess.find('2')==2:
                           end = 3
                        pass
                        if firsttimess.find('3')==2:
                           end = 4
                        pass
                        if firsttimess.find('4')==2:
                           end = 5
                        pass
                        if firsttimess.find('F')==2:
                           end = 6
                        pass
                        if firsttimess.find('5')==2:
                           end = 7
                        pass
                        if firsttimess.find('6')==2:
                           end = 8
                        pass
                        if firsttimess.find('7')==2:
                           end = 9
                        pass
                        if firsttimess.find('8')==2:
                           end = 10
                        pass
                        if firsttimess.find('9')==2:
                           end = 11
                        pass
                        if firsttimess.find('A')==2:
                           end = 12
                        pass
                        if firsttimess.find('B')==2:
                           end = 13
                        pass
                        if firsttimess.find('C')==2:
                           end = 14
                        pass
                        if firsttimess.find('D')==2:
                           end = 15
                        pass
                        while time <= end:
                           firstclass = firstnamess+'(' + firstteacherss +')  ' + firstwheress[0:len(firstwheress)-7] + ' ' + firstwheress[len(firstwheress)-7:len(firstwheress)]
                           sheet_1.cell(time,week).value = firstclass
                           sheet_1.cell(time,week).font = Font(size=8)
                           time += 1
                        pass
                    #第二個
                        if firstweekss.find('一')==2:
                           week = 7
                        pass
                        if firstweekss.find('二')==2:
                           week = 6
                        pass
                        if firstweekss.find('三')==2:
                           week = 5
                        pass
                        if firstweekss.find('四')==2:
                           week = 4
                        pass
                        if firstweekss.find('五')==2:
                           week = 3
                        pass
                        if firstweekss.find('六')==2:
                           week = 2
                        pass
                        if firstweekss.find('七')==2:
                           week = 1
                        pass
                        if firsttimess.find('1')==4:
                           time = 2
                           end = 2
                        pass
                        if firsttimess.find('2')==4:
                           time = 3
                           end = 3
                        pass
                        if firsttimess.find('3')==4:
                           time = 4
                           end = 4
                        pass
                        if firsttimess.find('4')==4:
                           time = 5
                           end = 5
                        pass
                        if firsttimess.find('F')==4:
                           time = 6
                           end = 6
                        pass
                        if firsttimess.find('5')==4:
                           time = 7
                           end = 7
                        pass
                        if firsttimess.find('6')==4:
                           time = 8
                           end = 8
                        pass
                        if firsttimess.find('7')==4:
                           time = 9
                           end = 9
                        pass
                        if firsttimess.find('8')==4:
                           time = 10
                           end = 10
                        pass
                        if firsttimess.find('9')==4:
                           time = 11
                           end = 11
                        pass
                        if firsttimess.find('A')==4:
                           time = 12
                           end = 12
                        pass
                        if firsttimess.find('B')==4:
                           time = 13
                           end = 13
                        pass
                        if firsttimess.find('C')==4:
                           time = 14
                           end = 14
                        pass
                        if firsttimess.find('D')==4:
                           time = 15
                           end = 15
                        pass
            
                        if firsttimess.find('1')==6:
                           end = 2
                        pass
                        if firsttimess.find('2')==6:
                           end = 3
                        pass
                        if firsttimess.find('3')==6:
                           end = 4
                        pass
                        if firsttimess.find('4')==6:
                           end = 5
                        pass
                        if firsttimess.find('F')==6:
                           end = 6
                        pass
                        if firsttimess.find('5')==6:
                           end = 7
                        pass
                        if firsttimess.find('6')==6:
                           end = 8
                        pass
                        if firsttimess.find('7')==6:
                           end = 9
                        pass
                        if firsttimess.find('8')==6:
                           end = 10
                        pass
                        if firsttimess.find('9')==6:
                           end = 11
                        pass
                        if firsttimess.find('A')==6:
                           end = 12
                        pass
                        if firsttimess.find('B')==6:
                           end = 13
                        pass
                        if firsttimess.find('C')==6:
                           end = 14
                        pass
                        if firsttimess.find('D')==6:
                           end = 15
                        pass
                        while time <= end:
                           firstclass = firstnamess+'(' + firstteacherss +')  ' + firstwheress[0:len(firstwheress)-7] + ' ' + firstwheress[len(firstwheress)-7:len(firstwheress)]
                           sheet_1.cell(time,week).value = firstclass
                           sheet_1.cell(time,week).font = Font(size=8)
                           time += 1
                        pass
                    except:
                        print('no')

                    if len(firstweekss) >= 4:
                        try:

                             if firstweekss.find('一')==4:
                                week = 7
                             pass
                             if firstweekss.find('二')==4:
                                week = 6
                             pass
                             if firstweekss.find('三')==4:
                                week = 5
                             pass
                             if firstweekss.find('四')==4:
                                week = 4
                             pass
                             if firstweekss.find('五')==4:
                                week = 3
                             pass
                             if firstweekss.find('六')==4:
                                week = 2
                             pass
                             if firstweekss.find('七')==4:
                                week = 1
                             pass
                             if firsttimess.find('1')==8:
                                time = 2
                                end = 2
                             pass
                             if firsttimess.find('2')==8:
                                time = 3
                                end = 3
                             pass
                             if firsttimess.find('3')==8:
                                time = 4
                                end = 4
                             pass
                             if firsttimess.find('4')==8:
                                time = 5
                                end = 5
                             pass
                             if firsttimess.find('F')==8:
                                time = 6
                                end = 6
                             pass
                             if firsttimess.find('5')==8:
                                time = 7
                                end = 7
                             pass
                             if firsttimess.find('6')==8:
                                time = 8
                                end = 8
                             pass
                             if firsttimess.find('7')==8:
                                time = 9
                                end = 9
                             pass
                             if firsttimess.find('8')==8:
                                time = 10
                                end = 10
                             pass
                             if firsttimess.find('9')==8:
                                time = 11
                                end = 11
                             pass
                             if firsttimess.find('A')==8:
                                time = 12
                                end = 12
                             pass
                             if firsttimess.find('B')==8:
                                time = 13
                                end = 13
                             pass
                             if firsttimess.find('C')==8:
                                time = 14
                                end = 14
                             pass
                             if firsttimess.find('D')==8:
                                time = 15
                                end = 15
                             pass
            
                             if firsttimess.find('1')==10:
                                end = 2
                             pass
                             if firsttimess.find('2')==10:
                                end = 3
                             pass
                             if firsttimess.find('3')==10:
                                end = 4
                             pass
                             if firsttimess.find('4')==10:
                                end = 5
                             pass
                             if firsttimess.find('F')==10:
                                end = 6
                             pass
                             if firsttimess.find('5')==10:
                                end = 7
                             pass
                             if firsttimess.find('6')==10:
                                end = 8
                             pass
                             if firsttimess.find('7')==10:
                                end = 9
                             pass
                             if firsttimess.find('8')==10:
                                end = 10
                             pass
                             if firsttimess.find('9')==10:
                                end = 11
                             pass
                             if firsttimess.find('A')==10:
                                end = 12
                             pass
                             if firsttimess.find('B')==10:
                                end = 13
                             pass
                             if firsttimess.find('C')==10:
                                end = 14
                             pass
                             if firsttimess.find('D')==10:
                                end = 15
                             pass
                             while time <= end:
                                firstclass = firstnamess+'(' + firstteacherss +')  ' + firstwheress[0:len(firstwheress)-7] + ' ' + firstwheress[len(firstwheress)-7:len(firstwheress)]
                                sheet_1.cell(time,week).value = firstclass
                                sheet_1.cell(time,week).font = Font(size=8)
                                time += 1
                             pass
                        except:
                            print('no')
                        



                else:
                    if firstweekss.find('一')==0:
                       week = 7
                    pass
                    if firstweekss.find('二')==0:
                       week = 6
                    pass
                    if firstweekss.find('三')==0:
                       week = 5
                    pass
                    if firstweekss.find('四')==0:
                       week = 4
                    pass
                    if firstweekss.find('五')==0:
                       week = 3
                    pass
                    if firstweekss.find('六')==0:
                       week = 2
                    pass
                    if firstweekss.find('七')==0:
                       week = 1
                    pass
                    if firsttimess.find('1')==0:
                       time = 2
                       end = 2
                    pass
                    if firsttimess.find('2')==0:
                       time = 3
                       end = 3
                    pass
                    if firsttimess.find('3')==0:
                       time = 4
                       end = 4
                    if firsttimess.find('4')==0:
                       time = 5
                       end = 5
                    pass
                    if firsttimess.find('F')==0:
                       time = 6
                       end = 6
                    pass
                    if firsttimess.find('5')==0:
                       time = 7
                       end = 7
                    pass
                    if firsttimess.find('6')==0:
                       time = 8
                       end = 8
                    pass
                    if firsttimess.find('7')==0:
                       time = 9
                       end = 9
                    pass
                    if firsttimess.find('8')==0:
                       time = 10
                       end = 10
                    pass
                    if firsttimess.find('9')==0:
                       time = 11
                       end = 11
                    pass
                    if firsttimess.find('A')==0:
                       time = 12
                       end = 12
                    pass
                    if firsttimess.find('B')==0:
                       time = 13
                       end = 13
                    pass
                    if firsttimess.find('C')==0:
                       time = 14
                       end = 14
                    pass
                    if firsttimess.find('D')==0:
                       time = 15
                       end = 15
                    pass
        
                    if firsttimess.find('1')==2:
                       end = 2
                    pass
                    if firsttimess.find('2')==2:
                       end = 3
                    pass
                    if firsttimess.find('3')==2:
                       end = 4
                    pass
                    if firsttimess.find('4')==2:
                       end = 5
                    pass
                    if firsttimess.find('F')==2:
                       end = 6
                    pass
                    if firsttimess.find('5')==2:
                       end = 7
                    pass
                    if firsttimess.find('6')==2:
                       end = 8
                    pass
                    if firsttimess.find('7')==2:
                       end = 9
                    pass
                    if firsttimess.find('8')==2:
                       end = 10
                    pass
                    if firsttimess.find('9')==2:
                       end = 11
                    pass
                    if firsttimess.find('A')==2:
                       end = 12
                    pass
                    if firsttimess.find('B')==2:
                       end = 13
                    pass
                    if firsttimess.find('C')==2:
                       end = 14
                    pass
                    if firsttimess.find('D')==2:
                       end = 15
                    pass
                    while time <= end:
                       firstclass = firstnamess+'(' + firstteacherss +')  ' + firstwheress[0:len(firstwheress)-7] + ' ' + firstwheress[len(firstwheress)-7:len(firstwheress)]
                       sheet_1.cell(time,week).value = firstclass
                       sheet_1.cell(time,week).font = Font(size=8)
                       time += 1
                    pass

                i += 1
            except:
                print('no')

        pass


        book_r.save('class.xlsx')
        reply = QMessageBox.warning(self,"完成","檔案輸出完成")
        driver.quit()

app = QApplication(sys.argv)
w = AppWindow()
w.show()
sys.exit(app.exec_())
